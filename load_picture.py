import requests
import io
from os import mkdir, path, remove, listdir
from selenium.webdriver.firefox.webdriver import WebDriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.firefox.options import Options
from PIL import Image
from openpyxl import load_workbook

options = Options()
options.add_argument('--headless')
service = Service(log_path=path.devnull, executable_path='.\Driver\geckodriver.exe')
driver = WebDriver(service=service, options=options)
s = requests.Session() #Use same session to download pictures

def create_folder():
    dir = 'Greentek kuvat'
    if not path.exists(dir):
        mkdir(dir)
    if len(listdir(dir)) != 0:
        for f in listdir(dir):
            remove(path.join(dir, f))

def get_web_urls(): # Get main links from Excel workbook and save product information
    wb = load_workbook("Greentek_linkit.xlsx")
    ws = wb.active
    for cell in ws.iter_rows(min_col=1,max_col=5, min_row=2):
        if cell[1].value is None:
            break
        article, carguments, dwld, yt_link = get_url_data(cell[1].value)
        cell[3].value = article
        cell[4].value = carguments
        if dwld:
            cell[0].value = 'x'
        if yt_link:
            cell[2].value = yt_link

    wb.save('Greentek_linkit.xlsx')

    # wb.close()

def get_url_data(url): # Get dynamically created image links and product information from web page
    image_urls = []
    yt_link = ''        # Youtube link if there is video available
    driver.get(url)
    WebDriverWait(driver, timeout=25).until(EC.visibility_of_element_located((By.CLASS_NAME,"thumbs")))
    elements = driver.find_elements(By.CLASS_NAME, 'thumbs-item')
    product = driver.find_element(By.CLASS_NAME, 'code').text.split()[-1]
    article = driver.find_element(By.TAG_NAME, 'article').get_attribute('innerText')
    carguments = driver.find_element(By.CLASS_NAME, 'carguments').get_attribute('innerText')

    for elem in elements:
        image_url = elem.value_of_css_property('background-image').strip('url"()').replace('thumb', 'original')

        if 'icon' in image_url:
            driver.find_element(By.CLASS_NAME, 'thumbs-video').click()
            WebDriverWait(driver, timeout=25).until(EC.visibility_of_element_located((By.TAG_NAME, "iframe")))
            yt_link = driver.find_element(By.TAG_NAME, "iframe").get_attribute('src')
        else:
            image_urls.append(image_url)

    dwld = download_images(image_urls, product)

    return article, carguments, dwld, yt_link

def download_images(image_urls:list, product:str): # Download images, check image size and save images
    i = 0
    dwld = True

    for image in image_urls:
        try:
            image_data = s.get(image, timeout=25)
            image_data.raise_for_status()
        except (requests.exceptions.RequestException) as e:
            print('Error: ', e)
            dwld = False
            continue

        print(image)
        imgs = Image.open(io.BytesIO(image_data.content))

        if imgs.width < imgs.height:
            imgs = imgs.resize((round(imgs.width*3), round(imgs.height*3)))
            imgs.thumbnail((800, 1100), resample=Image.LANCZOS)
        elif imgs.width >= imgs.height:
            imgs = imgs.resize((round(imgs.width*3), round(imgs.height*3)))
            imgs.thumbnail((1100, 800), resample=Image.LANCZOS)

        if imgs.mode in ('RGBA', 'P'):
            if imgs.mode in ('P'):
                imgs = imgs.convert('RGBA')
            bg = Image.new('RGBA', imgs.size, (255,255,255))
            imgs = Image.alpha_composite(bg, imgs).convert('RGB')

        imgs.save(f'Greentek kuvat/{product}_{(i+1)}.jpg', dpi=(300,300))

        i += 1
            
    return dwld # Return False if any of the images from image urls fails to download, else True

if __name__ == "__main__":
    create_folder()
    get_web_urls()
    driver.delete_all_cookies()
    driver.quit()
    s.close()